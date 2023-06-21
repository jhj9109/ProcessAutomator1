from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Side, Border, numbers
from openpyxl.utils import get_column_letter

from PIL import Image as PILImage
import json
import sys

# 두개이상의 파일에서 공통으로 가져가야할 규칙은 import해서 사용하기
from common_utils import get_worksheet_name, get_xlsx_file_name, load_json

thin_side_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

PX_TO_PT = 3 / 4
PT_TO_PX = 4 / 3 # 1.33

# Cell 한칸의 비율은 세로:가로 = 160:120 = 4:3
IMAGE_CELL_HEIGHT_PT = 160
IMAGE_CELL_HEIGHT_PX = IMAGE_CELL_HEIGHT_PT * PT_TO_PX

IMAGE_CELL_WIDTH_PT = 120
IMAGE_CELL_WIDTH_PT_6 = IMAGE_CELL_WIDTH_PT / 6

CELL_HEAD = 'A1'
MERGED_CELL_HEAD_RANGE = 'A1:D1'
HEAD_STRING = '세대별 부착 사진대지'

CELL_LABEL = 'A2'
CELL_VALUE = 'B2'
LABEL_STRING = '단지명:'

FIRST_ITEM_ROW_INDEX = 3

def set_common_head(ws, 단지명):
    # 1. 세대별 부착 사진대지
    ws.merge_cells(MERGED_CELL_HEAD_RANGE)
    ws[CELL_HEAD].value = HEAD_STRING
    
    # 2. 단지명: 해당 단지 이름
    ws[CELL_LABEL] = LABEL_STRING
    ws[CELL_VALUE] = 단지명

def set_item_format(ws, 몇동, 호수, r):
    # 이미지 들어갈 행 높이 조절
    ws.row_dimensions[r].height = IMAGE_CELL_HEIGHT_PT
    # 이미지 들어갈 셀 2개씩 병합
    ws.merge_cells(f'A{r}:B{r}')
    ws.merge_cells(f'C{r}:D{r}')
    # 이미지 아래 동 호수 입력
    ws[f'A{r+1}'].value = f'{몇동}동'
    ws[f'B{r+1}'].value = f'{호수}호'
    ws[f'C{r+1}'].value = f'{몇동}동'
    ws[f'D{r+1}'].value = f'{호수}호'

def set_items_format(ws, 몇동, 호수목록):
    row_index = FIRST_ITEM_ROW_INDEX
    for 호수 in 호수목록:
        set_item_format(ws, 몇동, 호수, row_index)
        row_index += 2

def set_whole_worksheet_style(ws):
    
    # 4개의 행 너비 조절
    ws.column_dimensions['A'].width = IMAGE_CELL_WIDTH_PT_6 / 2
    ws.column_dimensions['B'].width = IMAGE_CELL_WIDTH_PT_6 / 2
    ws.column_dimensions['C'].width = IMAGE_CELL_WIDTH_PT_6 / 2
    ws.column_dimensions['D'].width = IMAGE_CELL_WIDTH_PT_6 / 2
    
    # 스타일 설정
    font = Font(name='Arial', bold=True, italic=False)
    alignment = Alignment(horizontal='center', vertical='center')

    # 워크시트 전체에 스타일 적용
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment
            cell.border = thin_side_border

def create_new_seat(wb, 단지명, 몇동, 호수목록):
    
    worksheet_name = get_worksheet_name(단지명, 몇동)
    ws = wb.create_sheet(worksheet_name)
    
    set_common_head(ws, 단지명)
    set_items_format(ws, 몇동, 호수목록)
    set_whole_worksheet_style(ws)

    row_index = FIRST_ITEM_ROW_INDEX

    for _호수 in range(len(호수목록)):
        ws[f'A{row_index}'].border = Border(
            left=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        ws[f'B{row_index}'].border = Border(
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        ws[f'C{row_index}'].border = Border(
            left=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        ws[f'D{row_index}'].border = Border(
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        row_index += 2

def create_new_xlsx(아파트객체):
    # 워크북 생성
    wb = Workbook()

    단지명 = 아파트객체['단지명']
    동호수목록 = 아파트객체['동호수목록']
    # 대상세대수 = 아파트객체['대상세대수']
        
    # 각 동별 워크시트 생성
    for k, v in 동호수목록.items():
        몇동 = int(k) # josn에서의 객체의 key로서 문자열인 숫자인 타입이므로 인트형으로 변환
        호수목록 = v
        create_new_seat(wb, 단지명, 몇동, 호수목록)

    # 기본 생성된 워크시트 삭제
    wb.remove(wb.worksheets[0])

    # 워크북 저장
    save_filename = get_xlsx_file_name(단지명)
    wb.save(save_filename)
    wb.close()
    
'''
FORMAT_PERCENTAGE: '0%',
FORMAT_PERCENTAGE_00: '0.00%',
'''
def set_percent_style(cell):
    cell.number_format = numbers.FORMAT_PERCENTAGE
    
def set_text_alignment_center(cell):
    cell.alignment = Alignment(horizontal='center', vertical='center')

def get_longest_column_length(ws, min_row, max_row, col):
    return max(len(str(cell.value)) for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=col, max_col=col) for cell in row)

def set_column_width(ws, min_row, max_row, col):
    column_letter = get_column_letter(col)
    column_dimensions = ws.column_dimensions[column_letter]
    logest_column_length = get_longest_column_length(ws, min_row, max_row, col) # 2: 여유공간
    if col == 3:
        # 단지명
        column_dimensions.width = logest_column_length + 8
    else:
        column_dimensions.width = logest_column_length + 2
    

def create_new_summary_xlsx(아파트목록, 엑셀파일명):
    # 워크북 생성
    wb = Workbook()

    # 기본 워크시트 선택
    ws = wb.active

    # 추가
    ws['A1'].value = '순번'
    ws['B1'].value = '지역구'
    # 기존, 2개씩 밀림
    ws['C1'].value = '단지명'
    ws['D1'].value = '대상세대수'
    ws['E1'].value = '완료세대수'
    ws['F1'].value = '진행률'
    
    START_INDEX = 2
    
    for idx, 아파트객체 in enumerate(아파트목록):
        i = START_INDEX + idx
        # 추가
        ws[f'A{i}'].value = 아파트객체['순번']
        ws[f'B{i}'].value = 아파트객체['지역구명']
        # 기존, 2개씩 밀림
        ws[f'C{i}'].value = 아파트객체['단지명']
        ws[f'D{i}'].value = 아파트객체['대상세대수']
        ws[f'E{i}'].value = 0
        ws[f'F{i}'].value = f"=E{i} / D{i}"
        set_percent_style(ws[f'F{i}'])
    
    i = START_INDEX + len(아파트목록)
    ws.merge_cells(f'A{i}:C{i}')
    ws[f'A{i}'].value = "합계"
    ws[f'D{i}'].value = f"=SUM(D{START_INDEX}:D{i-1})"
    ws[f'E{i}'].value = f"=SUM(E{START_INDEX}:E{i-1})"
    ws[f'F{i}'].value = f"=E{i} / D{i}"
    set_percent_style(ws[f'F{i}'])
    
    # 스타일 적용
    for row in ws.iter_rows():
        for cell in row:
            set_text_alignment_center(cell)
    for col_num in range(1, 6+1):
        if col_num <= 3:
            set_column_width(ws, 1, i-1, col_num)
        else:
            set_column_width(ws, 1, i, col_num)

    # 워크북 저장
    wb.save(엑셀파일명)
    wb.close()

''' JSON 파일에서 config 읽기
{
    "아파트목록": [
        {
            "단지명": "서울번동3",
            "동호수목록": {
                "301": [ 101, 102, ..., 109 ]
            },
            "대상세대수" : 1234
        }
    ],
    "설정파일명": "아파트목록.json",
    "엑셀파일명": "요약.xlsx"
}
'''

'''
1. 아파트목록 정보가 담긴 json 파일 하나를 읽는다.
2. 각 아파트별로 하나의 엑셀파일을 만든다.
3. 하나의 엑셀 파일은 각 동별로 시트를 만든다.
'''
USE_DEFAULT_PATH = True

if __name__ == '__main__':

    try:
        if USE_DEFAULT_PATH:
            설정파일경로 = 'apartments.json'
        else:
            설정파일경로 = sys.argv[1]
    except IndexError:
        raise Exception(f"파일경로가 인자로 적절히 입력되지 않았습니다.")
    
    config = load_json(설정파일경로)
    
    아파트목록 = config['아파트목록']
    # 설정파일명 = config["설정파일명"]
    엑셀파일명 = config['엑셀파일명']

    # 각 아파트단지별 파일 하나 생성
    for 아파트객체 in 아파트목록:
        create_new_xlsx(아파트객체)
    
    # summary 파일 하나 생성
    create_new_summary_xlsx(아파트목록, 엑셀파일명)
    