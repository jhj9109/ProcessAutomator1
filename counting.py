from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from common_utils import load_json, get_xlsx_file_name, get_worksheet_name, get_apart_object

import sys
from pprint import pprint

from datetime import date

from colorama import Fore, Back, Style

FIRST_ITEM_ROW_INDEX = 3

data = {
    "서울번동3": '강북구',
    "서울번동5": '강북구',
    "서울번동2": '강북구',
    "서울가양": '강서구',
    "서울등촌9": '강서구',
    "서울등촌7": '강서구',
    "서울등촌1": '강서구',
    "서울등촌4": '강서구',
    "서울등촌6": '강서구',
    "서울등촌11": '강서구',
    "서울중계1": '노원구',
    "서울중계3": '노원구',
    "서울중계3(주거복지동)": '노원구',
    "서울중계9": '노원구',
    "서울중계9(주거복지동)": '노원구',
    "서울월계": '노원구',
    "서울오류": '노원구',
    "서울공릉": '노원구',
    "서울가좌": '마포구',
    "서울중구": '중구',
}

def insert_image_with_cell_height(ws, image_path, cell):
    img = Image(image_path)
    ws.add_image(img, cell.coordinate)

# def is_image_cell(ws, cell):
#     return any(image.anchor == cell.coordinate for image in ws._images)

def is_image_cell(image_obj, cell):
    return cell.coordinate in image_obj.anchors

def 사진체크하기(image_obj, ws, row):
    
    현관사진_cell = ws.cell(row=row, column=1)
    큐알사진_cell = ws.cell(row=row, column=3)
    
    return is_image_cell(image_obj, 현관사진_cell) and is_image_cell(image_obj, 큐알사진_cell)

def count_images(ws, image_obj, n):
    
    image_count = 0
    
    for idx in range(n):
        row_index = FIRST_ITEM_ROW_INDEX + 2 * idx
        if 사진체크하기(image_obj, ws, row_index):
            image_count += 1

    return image_count

def get_image_object(ws):
    image_obj = {
        images: ws._images,
        anchors: { img.anchor: img for img in ws._images },
    }
    return image_obj

def 워크시트_하나_조사하기(ws, n):
    
    image_obj = get_image_object(ws)
    
    cnt = count_images(ws, image_obj, n)

def 워크시트_유니크_사진수(ws):
    # count = len(set([ img.anchor for img in ws._images ]))
    count = len([ img for img in ws._images ])
    return count

def 워크시트_작업량(ws):
    return 워크시트_유니크_사진수(ws) // 2

def 진행률체크위한_사진체크하기(아파트객체, 엑셀파일명):
    
    단지명 = 아파트객체["단지명"]
    동호수목록 = 아파트객체["동호수목록"]
    대상세대수 = 아파트객체["대상세대수"]
    
    # 1. 타겟 파일 열기
    파일경로 = get_xlsx_file_name(단지명)
    wb = load_workbook(filename = 파일경로)
    
    # 2. 워크시트를 전체를 순회하면서, 각 시트에서 이미지 카운트 => 누적
    # 완료세대수 = sum([ 워크시트_하나_조사하기(wb[get_worksheet_name(단지명, 몇동)], len(호수목록)) for 몇동, 호수목록 in 동호수목록.items() ])
    완료세대수 = sum([ 워크시트_작업량(wb[get_worksheet_name(단지명, 몇동)]) for 몇동 in 동호수목록.keys() ])
    
    # 3. 열었던 타겟파일은 닫기
    wb.close()
    
    return 완료세대수

def 아파트단지_하나_카운팅(config, 단지명, 프린트여부 = True, 파일업데이트여부 = False):
    
    아파트목록 = config['아파트목록']
    # 설정파일명 = config["설정파일명"]
    엑셀파일명 = config['엑셀파일명']
    
    아파트객체 = get_apart_object(config, 단지명)

    완료세대수 = 진행률체크위한_사진체크하기(아파트객체, 엑셀파일명)
    대상세대수 = 아파트객체["대상세대수"]
    
    # 1. 프린트
    if 프린트여부:    
        진행률 = (완료세대수 / 대상세대수) * 100
        print(f"{단지명}: {완료세대수}호 / 전체 {대상세대수} 호 = {진행률}%")
    
    순번_열 = 1
    지역구_열 = 2
    단지명_열 = 3
    대상세대수_열 = 4
    완료세대수_열 = 5
    
    START_INDEX = 1
    
    # 2. 요약 파일 업데이트
    if 파일업데이트여부:
        wb = load_workbook(filename = 엑셀파일명)
        ws = wb.active
        
        cell = ws.cell(아파트객체["순번"] + 1, 완료세대수_열)
        cell.value = 완료세대수
        
        wb.save(엑셀파일명)
        wb.close()


def 카운팅_업데이트(config, 결과객체모음, 프린트여부 = True, 파일업데이트여부 = False):
    작업회차 = 2
    작업날짜 = "20230704" or date.today().strftime('%Y%m%d')
    
    # 아파트목록 = config['아파트목록']
    엑셀파일명 = config['엑셀파일명']

    for 단지명, 결과객체 in 결과객체모음.items():

        아파트객체 = get_apart_object(config, 단지명)

        작업량 = 결과객체["작업량"]
        신규_작업리스트 = 작업량["신규"]
        업데이트_작업리스트 = 작업량["업데이트"]

        완료세대수 = 작업량["기존"] + len(신규_작업리스트)
        대상세대수 = 아파트객체["대상세대수"]

        # pprint(신규_작업리스트)
        # pprint(업데이트_작업리스트)
        # print(f'작업량: 기존:{작업량["기존"]}, 신규:{len(신규_작업리스트)}, 업데이트:{len(업데이트_작업리스트)}')
    
        # 1. 프린트
        if 프린트여부:    
            진행률 = (완료세대수 / 대상세대수) * 100
            print(f"{Fore.RED}{단지명}{Style.RESET_ALL}: 기존:{작업량['기존']}, 신규{len(신규_작업리스트)}호, 업데이트{len(업데이트_작업리스트)}호 진행 => {완료세대수}호 / 전체 {대상세대수} 호 = {Fore.BLUE}{진행률:.0f}%{Style.RESET_ALL}")
        
        순번_열 = 1
        지역구_열 = 2
        단지명_열 = 3
        대상세대수_열 = 4
        완료세대수_열 = 5
        퍼센트계산_열 = 6
        
        START_INDEX = 1
        
        # 2. 요약 파일 업데이트
        if 파일업데이트여부:
            wb = load_workbook(filename = 엑셀파일명)
            ws = wb.active

            해당작업회차_기입할_열 = 퍼센트계산_열 + 작업회차

            # 0. 날짜 기입
            if not ws.cell(1, 해당작업회차_기입할_열).value:
                ws.cell(1, 해당작업회차_기입할_열).value = 작업날짜

            # 1. 완료세대수 업데이트
            완료세대수_cell = ws.cell(아파트객체["순번"] + 1, 완료세대수_열)
            완료세대수_cell.value = 완료세대수

            # 2. 오늘 작업량에 대한 업데이트
            기록_cell = ws.cell(아파트객체["순번"] + 1, 해당작업회차_기입할_열)
            기록_cell.value = f"신규{len(신규_작업리스트)}/업데이트{len(업데이트_작업리스트)}"
            
            wb.save(엑셀파일명)
            wb.close()

if __name__ == '__main__':

    config = load_json('./apartments.json')
    프린트여부 = True
    파일업데이트여부 = True
    for 단지명 in data.keys():
        try:
            아파트단지_하나_카운팅(config, 단지명, 프린트여부, 파일업데이트여부)
        except:
            print(f"{단지명} 카운팅 실패")
