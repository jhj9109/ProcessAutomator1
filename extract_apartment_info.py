from openpyxl import load_workbook
import json
from collections import defaultdict

from common_utils import load_json

def is_dong_ho(dong, ho):
    return isinstance(dong, int) and isinstance(ho, int)

def process_index_dong_ho(apartment_data, max_index, index, dong, ho):
    if is_dong_ho(dong, ho):
        apartment_data[dong].append(ho)
        if isinstance(index, int):
            max_index = max(max_index, index)
    return max_index

def 출력파일정보객체_만들기(sheetname, apartment_data, num_of_apartment):
    # 1325인 중계3까지는 하나의 파일로 할것이다.
    SPLIT_THRESHOLD = 1330

    출력파일정보객체 = dict()
    동리스트 = list(map(int, apartment_data.keys()))

    if num_of_apartment <= SPLIT_THRESHOLD:
        출력파일정보객체[sheetname] = 동리스트
    else:
        출력파일정보객체[sheetname+'_1'] = 동리스트[:len(동리스트)//2]
        출력파일정보객체[sheetname+'_2'] = 동리스트[len(동리스트)//2:]
    return 출력파일정보객체


def get_apartment_info(area_data, wb, sheetname, index):
    
    # 시트 열기, sheetname == 단지명
    ws = wb[sheetname]
    
    # 시트별로 동호수를 저장할 리스트 생성
    apartment_data = defaultdict(list)
    max_index = 0

    # 시트의 모든 행 순회
    for row in ws.iter_rows(min_row=3):
        # 동호수가 있는 열의 값을 조회 + 순번(처음부터 빈 집은 순번이 비어있을 수 있다! => 빼고 카운팅 됨)
        max_index = process_index_dong_ho(apartment_data, max_index, row[0].value, row[1].value, row[2].value)
        max_index = process_index_dong_ho(apartment_data, max_index, row[5].value, row[6].value, row[7].value)

    # 정렬하기
    for k, v in apartment_data.items():
        apartment_data[k] = sorted(v)
    
    apartment_info = {
        '단지명': sheetname,
        '동호수목록': apartment_data,
        '출력파일정보객체': 출력파일정보객체_만들기(sheetname, apartment_data, max_index),
        '대상세대수': max_index,
        '순번': index, # 1~20,
        '지역구명': area_data[sheetname],
    }

    return apartment_info

def get_new_data(apartment_data):
    result = dict()
    for k, v in apartment_data.items():
        lst = []
        i = 0
        while i < len(v):
            j = 1
            while i + j < len(v) and v[i]+j == v[i+j]:
                j += 1
            lst.append([v[i], v[i] + j - 1])
            i += j 
        result[k] = lst
    return result

def analyze_apartments(area_data, 추출할엑셀파일경로, 설정파일명, 엑셀파일명):

    # 엑셀 파일 열기
    wb = load_workbook(filename = 추출할엑셀파일경로)

    # 각 시트를 순회하여 아파트목록 데이터 생성
    아파트목록 = [get_apartment_info(area_data, wb, sheetname, i+1) for i, sheetname in enumerate(wb.sheetnames)]

    # JSON 파일로 저장
    config = {
        '아파트목록': 아파트목록,
        '설정파일명': 설정파일명,
        '엑셀파일명': 엑셀파일명,
    }

    with open(설정파일명, 'w', encoding='utf-8') as json_file:
        json.dump(config, json_file, ensure_ascii=False, indent=4)

''' 
1. excel 파일 하나에 단지별로 시트 하나에 명부가 적혀있다.
2. 각 시트별로 순회하며 단지 정보를 추출하여 하나의 아파트 객체를 만든다.
3. 만들어진 아파트 객체 리스트를 json 형식으로 파일로 저장한다.
'''
if __name__ == '__main__':
    
    DEFAULT_FILE_PATH = "./area.json"
    추출할엑셀파일경로 = '00 - 입주자서명부.xlsx'
    설정파일명 = 'apartments.json'
    엑셀파일명 = 'summary.xlsx'
    area_data = load_json(DEFAULT_FILE_PATH)

    # 아파트 분석 실행
    analyze_apartments(area_data, 추출할엑셀파일경로, 설정파일명, 엑셀파일명)
