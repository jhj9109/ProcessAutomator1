from collections import defaultdict
from openpyxl import Workbook, load_workbook
from common_utils import get_apart_object, get_worksheet_name, get_xlsx_file_name, load_json
import sys
import os
from data_processing_v1 import get_index_from_row_index

from image_utils import get_image_records
from datetime import date

from pprint import pprint


def 아파트_한_동_작업상태_업데이트하기(동호수별_작업상태, ws, 동, 호수목록):

    r = get_image_records(ws)

    for row_index in r.keys():
        index = get_index_from_row_index(row_index)
        호 = 호수목록[index]
        동호수별_작업상태[동][호] = '작업완료'


def 아파트_한_단지_작업상태_생성하기(base_path, 아파트객체, 단지명):

    동호수목록 = 아파트객체["동호수목록"]

    동호수별_작업상태 = {int(동): {호수: '' for 호수 in 호수목록} for 동, 호수목록 in 동호수목록.items()}

    # 엑셀이 쪼개져있든 하나든 >>>>> 하나의 시트는 동별로 되어있다.
    for 단지명2, 동목록 in 아파트객체["출력파일정보객체"].items():
        # 하나의 엑셀파일

        filename = os.path.join(base_path, get_xlsx_file_name(단지명2))
        wb = load_workbook(filename=filename)

        for 동 in 동목록:
            # 하나의 워크시트

            ws = wb[get_worksheet_name(단지명, 동)]
            호수목록 = 동호수목록[str(동)]

            아파트_한_동_작업상태_업데이트하기(동호수별_작업상태, ws, 동, 호수목록)

        wb.close()

    return 동호수별_작업상태


def 동호수별_작업상태_엑셀생성(동호수별_작업상태, 단지명):

    today_str = date.today().strftime('%Y%m%d')

    wb = Workbook()
    ws = wb.active

    동_컬럼 = 1
    호_컬럼 = 2
    작업상태_컬럼 = 3

    첫번째_아이템_로우 = 3

    ws.cell(1, 1).value = 단지명
    ws.cell(1, 2).value = today_str

    ws.cell(2, 동_컬럼).value = "동"
    ws.cell(2, 호_컬럼).value = "호"
    ws.cell(2, 작업상태_컬럼).value = "작업상태"

    i = 0
    count = defaultdict(int)

    for 동, 동객체, in 동호수별_작업상태.items():
        for 호, 작업상태 in 동객체.items():

            row = 첫번째_아이템_로우+i

            ws.cell(row, 동_컬럼).value = 동
            ws.cell(row, 호_컬럼).value = 호
            ws.cell(row, 작업상태_컬럼).value = 작업상태

            i += 1
            count[작업상태] += 1

    col = 1

    ws.cell(첫번째_아이템_로우+i, col).value = "총세대수"
    col += 1
    ws.cell(첫번째_아이템_로우+i, col).value = i+1
    col += 1

    for k, v in count.items():
        ws.cell(첫번째_아이템_로우+i, col).value = k
        col += 1
        ws.cell(첫번째_아이템_로우+i, col).value = v
        col += 1

    wb.save(filename=f"작업상태_{단지명}_{today_str}.xlsx")


def 아파트_한_단지_작업상태_추출하여_엑셀생성(단지명, base_path, config_path='./apartments.json'):

    config = load_json(config_path)
    아파트객체 = get_apart_object(config, 단지명)

    동호수별_작업상태 = 아파트_한_단지_작업상태_생성하기(base_path, 아파트객체, 단지명)
    동호수별_작업상태_엑셀생성(동호수별_작업상태, 단지명)


''' To use
python3 아파트_한_단지_작업상태_추출하여_엑셀생성.py "." "서울등촌7"
'''
if __name__ == '__main__':

    base_path = sys.argv[1]
    단지명 = sys.argv[2]

    아파트_한_단지_작업상태_추출하여_엑셀생성(단지명, base_path)
