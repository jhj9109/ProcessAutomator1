# for openpyxl, 엑셀류를 다루는 표준을 따르는 라이브러리, Image를 다룰때 내부적으로 pillow를 사용해서 인스톨 필수
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
# from PIL import Image as PILImage # 이미지 크기를 줄이고 싶을때, 이것을 활용해야할듯.

import json  # json 파일에서 설정을 사용하고자함
import sys  # 파일 실행시 인자 받아서 활용하기
import os  # 해당경로 하위 파일 파악하는데 활용
import re  # 파일명에서 정보 추출하는데 활용

from collections import defaultdict

# 두개이상의 파일에서 공통으로 가져가야할 규칙은 import해서 사용하기
from common_utils import get_worksheet_name, get_xlsx_file_name, sorted_file_entries, load_json, get_apart_object
from image_preprocessing import handle_rotated_or_mpo_image
from image_utils import get_image_records

import pdb

from colorama import Fore, Back, Style

from constants import IMAGE_CELL_HEIGHT_PX, IMAGE_CELL_WIDTH_PX, FIRST_ITEM_ROW_INDEX

DEBUG_MODE = False

CELL_HEAD = 'A1'
MERGED_CELL_HEAD_RANGE = 'A1:D1'
HEAD_STRING = '세대별 부착 사진대지'

CELL_LABEL = 'A2'
CELL_VALUE = 'B2'
LABEL_STRING = '단지명:'

ENUM_현관사진 = 0
ENUM_큐알사진 = 1

신규_플래그 = 0
업데이트_플래그 = 1


def insert_image_with_cell_height(ws, image_path, cell, 현관사진여부, 기존이미지인덱스=-1):

    handle_rotated_or_mpo_image(image_path)

    img = Image(image_path)

    # 이미지 사이즈 조절 => 원본 크기를 조절하는것은 아니였던듯? 하려면 pillow의 Image 직접 사용해야하는듯
    img.width = IMAGE_CELL_HEIGHT_PX * img.width / img.height
    img.height = IMAGE_CELL_HEIGHT_PX

    if 현관사진여부 and img.width > img.height:
        img.height = img.height * IMAGE_CELL_WIDTH_PX / img.width
        img.width = IMAGE_CELL_WIDTH_PX

    # 이미지 삽입 or 교체
    if 기존이미지인덱스 != -1:
        # print(f"{기존이미지인덱스}: 이미지교체")
        img.anchor = cell.coordinate
        ws._images[기존이미지인덱스] = img
    else:
        # print(f"{기존이미지인덱스}: 이미지추가")
        ws.add_image(img, cell.coordinate)


def check_extension(filename, extname):
    pattern = rf"\.{extname}$"
    return bool(re.search(pattern, filename))


def extract_extension(filename):
    pattern = r"\.([^.]+)$"
    match = re.search(pattern, filename)
    if match:
        return match.group(1)
    else:
        return None


def extract_filename_components(filename):
    pattern = r"^(.+)\.([^.]+)$"
    match = re.match(pattern, filename)
    if match:
        basename, extension = match.groups("")
        return basename, extension
    else:
        return "", "", ""


def 동호수_유효성체크(동호수목록, 동, 호수):
    try:
        return 동호수목록[동].index(int(호수))
    except KeyError:
        return -1
    except ValueError:
        return -1


def get_row_index(index):
    return FIRST_ITEM_ROW_INDEX + index * 2


def 워크시트하나에_이미지_한개_삽입하기(ws, r, 현관사진_파일경로, 큐알사진_파일경로, 인덱스):

    result = 신규_플래그

    현관사진_COLUMN = 1
    큐알사진_COLUMN = 3

    row_index = get_row_index(인덱스)

    현관사진_cell = ws.cell(row=row_index, column=1)
    큐알사진_cell = ws.cell(row=row_index, column=3)

    index = [-1, -1]

    # pdb.set_trace()

    if row_index in r:
        # print("업데이트")
        index = [r[row_index][현관사진_COLUMN], r[row_index][큐알사진_COLUMN]]
        result = 업데이트_플래그

    insert_image_with_cell_height(ws, 현관사진_파일경로, 현관사진_cell, True, index[0])
    insert_image_with_cell_height(ws, 큐알사진_파일경로, 큐알사진_cell, False, index[1])

    return result


def 동_호수_큐알사진여부_패턴매칭하기(filename):
    # pattern1 = r'^(\d+)동\s*(\d+)호\s*(\(1\))?(\(2\))?\s*\.(?:png|jpg|jpeg|JPEG|PNG|JPG)$'
    # pattern2 = r'^(\d+)동\s*(\d+)호\s*(\(1\))?(\(2\))?\s*\.(?:png|jpg|jpeg|JPEG|PNG|JPG)$'
    # pattern3 = r'^(\d+)[^\d]+(\d+)[^\d]+(\(1\))?(\(2\))?\s*\.(?:png|jpg|jpeg|JPEG|PNG|JPG)$'
    # pattern4 = r'^(\d+)[^\d\(\)]+(\d+)[^\d\(\)]+(\(1\))?(\(2\))?\s*\.(?:png|jpg|jpeg|JPEG|PNG|JPG)$'
    # (1), (2) 반드시
    pattern5 = r'^(\d+)[^\d\(\)]+(\d+)[^\d\(\)]+\(([12])\)\s*\.(?:png|jpg|jpeg|JPEG|PNG|JPG)$'
    pattern6 = r'^(\d+)-(\d+)(_1)?\.(?:png|jpg|jpeg|JPEG|PNG|JPG)$'

    # 111-444.jpg => 현관사진, 111-444_1.jpg => 큐알사진
    matched = re.match(pattern6, filename)
    if matched:
        동, 호수, 큐알사진여부 = matched.groups(False)
        큐알사진여부 = bool(큐알사진여부)
        # print(f"{Fore.RED}{filename}{Style.RESET_ALL}")
        return (동, 호수, 큐알사진여부)

    # 111     444    (1)   .jpg, 111     444    (2)   .jpg
    matched = re.match(pattern5, filename)
    if matched:
        동, 호수, 큐알사진여부 = matched.groups(False)
        큐알사진여부 = 큐알사진여부 == '2'
        # print(f"{Fore.BLUE}{filename}{Style.RESET_ALL}")
        return (동, 호수, 큐알사진여부)

    # print(f"{Fore.YELLOW}{filename}{Style.RESET_ALL}")
    return None


def 동별_작업분_생성하기(동호수목록, file_entries):
    # 동별로 하나의 시트 => 동별로 구분 짓기.
    동별_작업분 = {int(동): defaultdict(lambda: ['', '', -1]) for 동 in 동호수목록.keys()}
    ''' 예시
    동별_작업분[101][777] = [현관사진경로, 큐알사진경로, 워크시트명, 호수 - 시작호수]
    - 워크시트명과 호수-시작호수는 필요 없어진것일수도 있지만.... 일단 냅두자
    '''

    파일명이_매칭이_안되어_실패, 유효성체크_실패 = [], []

    for entry in file_entries:
        filename = entry.name

        result = 동_호수_큐알사진여부_패턴매칭하기(filename)

        if result is None:
            파일명이_매칭이_안되어_실패.append(filename)
            continue

        동, 호수, 큐알사진여부 = result
        인덱스 = 동호수_유효성체크(동호수목록, 동, 호수)

        if 인덱스 == -1:
            유효성체크_실패.append(filename)
            continue

        동, 호수 = map(int, [동, 호수])
        사진모드 = ENUM_큐알사진 if 큐알사진여부 else ENUM_현관사진

        동별_작업분[동][호수][사진모드] = entry.path
        동별_작업분[동][호수][2] = 인덱스

    # # 디버깅중
    # for i, lst in enumerate([파일명이_매칭이_안되어_실패, 유효성체크_실패]):
    #     print(["파일명이_매칭이_안되어_실패", "유효성체크_실패"][i])
    #     print(lst)

    return 동별_작업분, 파일명이_매칭이_안되어_실패, 유효성체크_실패


def 작업분_기존엑셀에_반영하기(base_path, 출력파일정보객체, 단지명, 동호수목록, file_entries):

    동별세이브_플래그 = False
    호별세이브_플래그 = False

    동별_작업분, 파일명이_매칭이_안되어_실패, 유효성체크_실패 = 동별_작업분_생성하기(동호수목록, file_entries)
    성공적으로업데이트, 사진_하나라도_없어_실패 = [], []
    작업량 = {"기존": 0, "신규": [], "업데이트": []}  # 신규, 업데이트

    for 출력파일명, 동목록 in 출력파일정보객체.items():  # "서울중계9_1", [901, 902, 903, 904, 905]

        파일경로 = os.path.join(base_path, get_xlsx_file_name(출력파일명))
        wb = load_workbook(filename=파일경로)

        for 동 in 동목록:

            워크시트명 = get_worksheet_name(단지명, 동)
            ws = wb[워크시트명]
            r = get_image_records(ws)
            # print(r)
            작업량["기존"] += len(r)

            호별작업물 = 동별_작업분[동]

            for 현관사진_파일경로, 큐알사진_파일경로, 인덱스 in 호별작업물.values():

                if 현관사진_파일경로 != '' and 큐알사진_파일경로 != '':

                    result = 워크시트하나에_이미지_한개_삽입하기(
                        ws, r, 현관사진_파일경로, 큐알사진_파일경로, 인덱스)
                    if result == 신규_플래그:
                        작업량["신규"].append({
                            "현관사진_파일경로": 현관사진_파일경로,
                            "큐알사진_파일경로": 큐알사진_파일경로,
                            "인덱스": 인덱스
                        })
                    else:
                        작업량["업데이트"].append({
                            "현관사진_파일경로": 현관사진_파일경로,
                            "큐알사진_파일경로": 큐알사진_파일경로,
                            "인덱스": 인덱스
                        })
                    성공적으로업데이트.append((현관사진_파일경로, 큐알사진_파일경로, 인덱스))
                else:
                    사진_하나라도_없어_실패.append((현관사진_파일경로, 큐알사진_파일경로, 인덱스))

                if 호별세이브_플래그:
                    wb.save(파일경로)
                    wb.close()
                    print(f"호별 세이브: {현관사진_파일경로}, {큐알사진_파일경로} 완료")
                    wb = load_workbook(filename=파일경로)

            if not 호별세이브_플래그 and 동별세이브_플래그:
                wb.save(파일경로)
                wb.close()
                print(f"{단지명} {동}동: 동별 세이브 완료")
                wb = load_workbook(filename=파일경로)

        if not 호별세이브_플래그 and not 동별세이브_플래그:
            wb.save(파일경로)
            wb.close()
            print(f"{출력파일명}: 파일별 세이브 완료")
    결과 = {
        "작업량": 작업량,
        # "성공적으로업데이트": 성공적으로업데이트, # 작업량과 중복
        "사진_하나라도_없어_실패": 사진_하나라도_없어_실패,
        "파일명이_매칭이_안되어_실패": 파일명이_매칭이_안되어_실패,
        "유효성체크_실패": 유효성체크_실패
    }
    return 결과
    # # 디버깅중
    # for i, lst in enumerate([성공적으로업데이트, 사진_하나라도_없어_실패]):
    #     print(["성공적으로업데이트", "사진_하나라도_없어_실패"][i])
    #     print(lst)


DEFAULT_CONFIG_FILE_PATH = "./apartments.json"


def update_one_apartment(config, folder_path, 단지명, base_path):

    # 1. 유효한 아파트 단지인지 체크
    아파트객체 = get_apart_object(config, 단지명)

    # 2. 폴더 하위 모든 파일 추출 => path까지 가진 DirEntry로 변경
    file_entries = sorted_file_entries(-1, folder_path)

    # 2-1. 하나도 파일이 없으면 에러
    if not file_entries:
        raise Exception(f"{folder_path}디렉토리 아래에 파일이 존재하지 않습니다.")

    # 3. 아파트객체에서 동호수목록을 가지고 유효성체크하며 기존 엑셀에 반영
    동호수목록 = 아파트객체["동호수목록"]

    # 4. 엑셀 파일 열기 위한 정보. json파일로 미리 만들어 놓던지 한다.
    출력파일정보객체 = 아파트객체["출력파일정보객체"]
    결과객체 = 작업분_기존엑셀에_반영하기(base_path, 출력파일정보객체, 단지명, 동호수목록, file_entries)
    return 결과객체


'''
1. 먼저 설정파일을 읽어서 아파트목록 데이터를 취한다.
2. 반영할 폴더명 & 단지명을 커맨드라인인수로 입력 받는다.
3. 단지명의 유효성 체크후, 유효하면 해당 아파트객체와 해당 아파트의 엑셀파일을 준비한다.
4. 폴더를 순회하여, 모든 파일에 대하여 엑셀파일에 삽입할 준비를 한다. => 덮어쓰기모드?
5. 업데이트된 엑셀파일을 저장한다.
'''
if __name__ == '__main__':

    config = load_json(DEFAULT_CONFIG_FILE_PATH)

    # 0. 커맨드라인 인수로부터 정보 입력 받음
    folder_path = sys.argv[1]
    단지명 = sys.argv[2]
    base_path = "./" if len(sys.argv) < 4 else sys.argv[3]

    update_one_apartment(config, folder_path, 단지명, base_path)
